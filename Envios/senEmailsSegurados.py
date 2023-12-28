import os
import openpyxl
import win32com.client as win32
from tkinter import filedialog




# Get the current working directory
cwd = os.getcwd()

baseanalista = filedialog.askopenfilename(title='ENCONTRAR A PLANILHA COM OS E-MAILS')

# Load the Excel workbook
workbook = openpyxl.load_workbook(os.path.join(baseanalista))

# Select the sheet
sheet = workbook["Plan1"]

# Get the Outlook application object
outlook = win32.Dispatch('Outlook.Application')


# Iterate through the rows in the sheet
for i in range(2, sheet.max_row + 1):

    numeropasta = sheet.cell(row=i, column=2).value
    emailsegurado = sheet.cell(row=i, column=4).value
    nsinistro = sheet.cell(row=i, column=5).value
    datasinistro = sheet.cell(row=i, column=6).value
    placasegurado = sheet.cell(row=i, column=7).value
    modelocarro = sheet.cell(row=i, column=8).value
    
    # Get the recipient name
    #recipient_name = sheet.cell(row=i, column=5).value

    # Get the recipient email address
    recipient_email = sheet.cell(row=i, column=4).value

    # Get the CC email address
    cc_email = sheet.cell(row=i, column=9).value
  
    # Create a new email
    mail = outlook.CreateItem(0)

    # Set the recipient and CC email addresses
    mail.To = recipient_email
    mail.BCC = cc_email

   

    # Set the email subject
    mail.Subject = f"Documentação {datasinistro}"

    # Set the email text
    mail.Body = f"Olá, tudo bem?\n\nTrabalho na DEBT, escritório prestador de serviços para a HDI Seguros.\n\nEstou enviando este e-mail para pedir uma ajuda na localização do causador do acidente que ocorreu em {datasinistro}, envolvendo o veículo {modelocarro}- {placasegurado}.\n\nO número do sinistro é: {nsinistro}\n\nEm toda documentação fornecida à HDI Seguros não constam os dados do causador do acidente.\n\nConseguiria nos ajudar com alguma das informações abaixo:\n\n* Nome ou CPF do causador do acidente;\n* Telefone do causador;\n* Placa do veículo do causador;\n* Fotos do acidente;\n* Boletim de ocorrência. \n\nPedimos por gentileza que nos encaminhe os dados e documentos em resposta a este e-mail.\n\nCaso tenha alguma dúvida converse com seu corretor.\n\nConto com a sua ajuda para resolvermos essa pendência o quanto antes.\n\nSe precisar de qualquer informação adicional pode nos contatar através desse e-mail, pelo whatsapp nº wa.me/554333785462 ou tel: 0800 400 5450.\n\nObs: Caso este e-mail tenha sido enviado ao corretor é porque o mesmo foi cadastro como e-mail do segurado no portal da HDI Seguros. Da mesma forma pedimos a gentileza de nos auxiliar na obtenção das informações necessárias.\n\nNúmero de referência: {numeropasta} \n\nDesde já muito obrigado!"

    # Open the email in Outlook
    mail.Display()
    
# close all opened objects
workbook.close()







#Olá, tudo bem?\n\nTrabalho na DEBT, escritório prestador de serviços para a HDI Seguros.\n\nEstou enviando este e-mail para pedir uma ajuda na localização do causador do acidente que ocorreu em {datasinistro}, envolvendo o veículo {modelocarro}- {placasegurado}.\n\nO número do sinistro é: {nsinistro}\n\nEm toda documentação fornecida à HDI Seguros não constam os dados do causador do acidente.\n\nConseguiria nos ajudar com alguma das informações abaixo:\n\n* Nome ou CPF do causador do acidente;\n* Telefone do causador;\n* Placa do veículo do causador;\n* Fotos do acidente;\n* Boletim de ocorrência. \n\nPedimos por gentileza que nos encaminhe os dados e documentos em resposta a este e-mail.\n\nCaso tenha alguma dúvida converse com seu corretor.\n\nConto com a sua ajuda para resolvermos essa pendência o quanto antes.\n\nSe precisar de qualquer informação adicional pode nos contatar através desse e-mail, pelo whatsapp nº wa.me/554333785462 ou tel: 0800 400 5450.\n\nObs: Caso este e-mail tenha sido enviado ao corretor é porque o mesmo foi cadastro como e-mail do segurado no portal da HDI Seguros. Da mesma forma pedimos a gentileza de nos auxiliar na obtenção das informações necessárias.\n\nNúmero de referência: {numeropasta} \n\nDesde já muito obrigado!