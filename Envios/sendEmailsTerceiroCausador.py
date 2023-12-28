import os
import openpyxl
import win32com.client
from tkinter import filedialog
from win32com.client import Dispatch
import win32api
import pythoncom



# Get the current working directory
cwd = os.getcwd()

baseanalista = filedialog.askopenfilename(title='ENCONTRAR A PLANILHA COM OS E-MAILS')


# Load the Excel workbook
workbook = openpyxl.load_workbook(os.path.join(baseanalista))

# Select the sheet
sheet = workbook["Plan1"]

# Get the Outlook application object
outlook = win32com.client.Dispatch('outlook.application')


# Iterate through the rows in the sheet
for i in range(2, sheet.max_row + 1):

    numeropasta = sheet.cell(row=i, column=1).value
    emailterceiro = sheet.cell(row=i, column=3).value
    datasinistro = sheet.cell(row=i, column=5).value
    nometerceiro = sheet.cell(row=i, column=4).value
    nomefuncionario = sheet.cell(row=i, column=6).value
        
    # Get the recipient name
    #recipient_name = sheet.cell(row=i, column=5).value

    # Get the recipient email address (corretor)
    recipient_email = sheet.cell(row=i, column=3).value

    # Get the CC email address (e-mail pipedrive)
    cc_email = sheet.cell(row=i, column=8).value
  
    # Create a new email
    mail = outlook.CreateItem(0)

    # Set the recipient and CC email addresses
    mail.To = recipient_email
    mail.BCC = cc_email

   

    # Set the email subject
    mail.Subject = f"Documentação {datasinistro}"

    # Set the email text
    mail.Body = f"Olá\n\n{nometerceiro}\n\n!\n\nTudo bem?\n\nMe chamo {nomefuncionario}, presto serviços para a HDI Seguros S/A.\n\nEstou entrando em contato para tratar dos prejuízos decorrentes do acidente de trânsito ocorrido na data de\n{datasinistro}\nenvolvendo a sua pessoa.\n\nDe acordo com a documentação que possuímos, você é o responsável pelos prejuízos causados ao veículo segurado pela HDI Seguros S/A.\n\nA seguradora realizou o conserto do veículo segurado e agora busca o ressarcimento dos valores gastos já descontado o valor da franquia.\n\nEssa cobrança é autorizada por lei e reconhecida por todos os tribunais (súmula 188 do STF, artigo 786 e 934 do Código Civil), caso tenha dúvida consulte um advogado de confiança.\n\nAinda estamos em esfera amigável, passível de negociação do valor a ser ressarcido, evitando assim maiores transtornos com eventual demanda judicial.\n\nSe seu veículo tinha seguro na data do ocorrido, nos informe qual era a Seguradora e o número do sinistro que resolvemos diretamente com a sua seguradora.\n\nCaso não tenha seguro ou tenha alguma dúvida, entre em contato pelo telefone 0800 400 5450, pelo whatsapp (43) 3378-5450 ou mesmo respondendo este e-mail.\n\nPor fim, é importante frisar que muitas pessoas na mesma situação já realizaram acordo amigável e não se arrependeram.\n\nNós temos um prazo restrito para realizar acordo no âmbito amigável.\n\nAproveite a oportunidade e faça uma proposta de pagamento antes que o processo seja encaminhado aos advogados para ajuizamento.\n\nFico no aguardo de seu retorno.\n\nP.P. HDI Seguros\n\nFicamos à disposição para esclarecer quaisquer dúvidas que tiver.\n\nNúmero de referência:\n{numeropasta}\n."

    # Open the email in Outlook
    mail.Display()
    
# close all opened objects
workbook.close()

