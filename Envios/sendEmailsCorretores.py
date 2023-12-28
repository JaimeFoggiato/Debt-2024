import os
import openpyxl
import win32com.client
from tkinter import filedialog
from win32com.client import Dispatch
import win32api
import pythoncom



# Get the current working directory
cwd = os.getcwd()

baseanalista = filedialog.askopenfilename(title='ENCONTRAR A PLANILHA COM OS E-MAILS')


# Load the Excel workbook
workbook = openpyxl.load_workbook(os.path.join(baseanalista))

# Select the sheet
sheet = workbook["Plan1"]

# Get the Outlook application object
outlook = win32com.client.Dispatch('outlook.application')


# Iterate through the rows in the sheet
for i in range(2, sheet.max_row + 1):

    numeropasta = sheet.cell(row=i, column=2).value
    emailsegurado = sheet.cell(row=i, column=4).value
    nsinistro = sheet.cell(row=i, column=5).value
    datasinistro = sheet.cell(row=i, column=6).value
    placasegurado = sheet.cell(row=i, column=7).value
    modelocarro = sheet.cell(row=i, column=8).value
    
    # Get the recipient name
    #recipient_name = sheet.cell(row=i, column=5).value

    # Get the recipient email address (corretor)
    recipient_email = sheet.cell(row=i, column=10).value

    # Get the CC email address (e-mail pipedrive)
    cc_email = sheet.cell(row=i, column=9).value
  
    # Create a new email
    mail = outlook.CreateItem(0)

    # Set the recipient and CC email addresses
    mail.To = recipient_email
    mail.BCC = cc_email

   

    # Set the email subject
    mail.Subject = f"Documentação {datasinistro}"

    # Set the email text
    mail.Body = f"Olá, tudo bem?\n\nTrabalho na DEBT, escritório prestador de serviços para a HDI Seguros.\n\nEstamos entrando em contato para pedir uma ajuda na localização do causador do acidente que ocorreu em {datasinistro}, envolvendo o veículo {modelocarro}- {placasegurado}.\n\nO número do sinistro é: {nsinistro}\n\nEstamos tentando localizar o causador do acidente mas não estamos tendo sucesso.\n\nVerificando o sinistro no portal da HDI, não encontramos dados ou documentos que apontem informações sobre o causador do acidente.\n\nConseguiria nos ajudar com alguma das informações abaixo:\n\n* Nome ou CPF do causador do acidente;\n* Telefone do causador;\n* Placa do veículo do causador;\n* Fotos do acidente;\n* Boletim de ocorrência.\n\nCom a localização do terceiro e realização de acordo para ressarcimento dos valores gastos pela companhia, os custos com a sinistralidade da HDI Seguros reduzirão e por isso sua ajuda é muito importante.\n\nEm consequência, com os custos de sinistralidade mais baixos, haverá também possibilidade de melhores condições nas negociações dos seguros pelos corretores no médio prazo, facilitando nas questões comerciais.\n\nOu seja, todos ganham com a localização do terceiro causador do acidente.\n\nEntendemos que essa parceria entre corretores e Seguradora é muito importante para ajudar a companhia na busca do ressarcimento.\n\nCaso já tenha anexado os dados ou documentos recentemente no portal da HDI, pedimos escusas e que desconsidere esse e-mail.\n\nSe precisar de qualquer informação adicional pode nos contatar através desse e-mail, pelo whatsapp nº wa.me/554333785462 ou tel: 0800 400 5450.\n\nNúmero de referência:\n\n {numeropasta}\n\n\n\n Desde já agradecemos pela ajuda."

    # Open the email in Outlook
    mail.Display()
    
# close all opened objects
workbook.close()



