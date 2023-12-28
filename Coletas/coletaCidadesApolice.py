from cgitb import text
from email import header
from gettext import gettext
from tracemalloc import stop
from turtle import goto
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup as BS
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import csv
import time
import pandas
from openpyxl import load_workbook
from openpyxl import workbook
import requests
from bs4 import BeautifulSoup
import re  #instalar

print('Abrindo o Chrome...')
#driver = Chrome(executable_path='C:\WebDriver\bin\chromedriver.exe')
#driver = webdriver.Chrome(ChromeDriverManager().install())

driver = webdriver.Chrome()

#driver.get("https://www.hdi.com.br/hdiprestador/") talvez seja o erro


print('Acessando o Site da HDI...')
driver.get("https://www.hdi.com.br/hdiprestador/")
wait = WebDriverWait(driver,20,poll_frequency=1)

print('usuário...')
driver.find_element(By.ID, 'm_prestserv').send_keys('debt')
print('senha...')
#driver.find_element_by_id('doc').click()
driver.find_element(By.ID,'m_senha').send_keys('De20bet23*')
print('prestador...')
driver.find_element(By.XPATH, '//*[@id="m_prest_oficina"]/option[3]').click()
print('entrando')
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="login_prestador"]/div[2]/button').click()
time.sleep(1)
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="bt_pesquisar"]')))
print('loguei!')
time.sleep(1)

       
# como pegar as informaçõe da sessão do selenium e transformar em request
cookies1 = {}

selenium_cookies = driver.get_cookies()

for cookie in selenium_cookies:
    cookies1[cookie['name']] = cookie['value']

urlprincipal = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/dsp_cad_processo_tp2_scp.htm'

urlpesquisa = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/dsp_pesq_proc_v02_scp.htm'

urlsinistro = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/mrsicc03x.htm'

urlapolice = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/hdidigital/act_documento_auto_digital.htm'


Headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:45.0) Gecko/20100101 Firefox/45.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Charset": "ISO-8859-1,utf-8;q=0.7,*;q=0.3",
            "Accept-Encoding": "none",
            "Accept-Language": "en-US,en;q=0.8",
            "Ajax-Response": "true",
            "Connection": "keep-alive",        
}


planilha = pandas.read_excel(r'C:\Users\Jaime Foggiato\Desktop\ProjetosDEBT-main\planilhaEmails.xlsx', engine='openpyxl') #não esquecer de mudar o diretório, toda vez que for rodar em um pc novo.

linha = 0

valor_nao_encontrado = "N/A"


while linha != len(planilha.index):  

    if (planilha['AUX'][linha]) == 0: 

        
        
        pasta = int(planilha['PASTA'][linha])

        
        
        paramsprincipal = {
            "sh": "clljlkkdaidllcBd",
            "us": "debt",
            "senha": "",
            "m_us": "debt",
            "m_municipio": "",
            "m_cod_processo": pasta,
            "m_recid": "",
            "dir": "pesquisar",
            "m_chamado_ressarc_judi": "", 
            "m_chamado_ressarc_amig": "",
            "enc_us": "ZGVidA==",
        
        }
        
        
        

        

        paginaprincipal = requests.post(urlprincipal, data=paramsprincipal, headers=Headers, cookies=cookies1)

        doc = BeautifulSoup(paginaprincipal.text, "html.parser")
        
        nsinistro_element = doc.find(id='m_num_sinistro')
        
        if nsinistro_element:
            nsinistro = nsinistro_element.get('value', 'N/A')
        else:
            nsinistro = 'Elemento não encontrado'
            linha += 1
            continue

        #nsinistro = doc.find(id='m_num_sinistro')['value']
        
            
        nsinistro = nsinistro[len(nsinistro)-15:]

        datasinistro = doc.find(id='m_dat_sinistro')['value']
        

        
        

        paramspesquisa = {
            "m_operacao": "",
            "us": "debt",
            "sh: saEekbalDiitlLbl": "",
            "m_pesquisa": "outros",
            "m_cod_processo": "", 
            "progr_ant": "act_login_scp.htm",
            "dir": "pesquisar",
            "m_disabled": "",
            "m_portal_prest": "SIM",
            "m_prestserv": "DEBT",
            "m_senha": "De20bet23*",
            "m_user_id": "debt",
            "m_cod_tipusu": "011",
            "enc_us": "ZGVidA==",
            "m_cod_empresa": "01",
            "m_nom_autor": "",
            "m_cod_sucursal": "", 
            "m_nom_reu": "",
            "m_cod_carteira": "", 
            "m_ca": "9934",
            "m_cod_ramo": "",
            "m_pasta": "",
            "m_num_sinistro": "", 
            "m_cod_cad_processo": "",
            "m_sinuss": "", 
            "m_tributario": "",
            "m_trabalhista": "",
            "m_susep_procon": "",
        
        }
        
        paginapesquisa = requests.post(urlpesquisa, data=paramspesquisa, headers=Headers, cookies=cookies1)
        
        docpesquisa = BeautifulSoup(paginapesquisa.text, "html.parser")
        
        token = docpesquisa.find('script', {'JavaScript': 'tokenSec'})
        
        print(token)
        
                  
    


        paramssinistro = {
            "isRevamp": "",
            "m_cor_padrao": "",
            "m_img_padrao": "",
            "m_tit_padrao": "",
            "sel_issinimport": nsinistro,
            "m_login": "debt",
            "ag_int": "0",
            "m_sinuss": nsinistro,
            "seg_cons": "" ,
            "hid_chamou": "ADVG",
            "hid_pendencia": "",
            "m_seq_envio": "",
            "cad_avaria_prv": "",
        
        }


        paginasinistro = requests.post(urlsinistro, data=paramssinistro, headers=Headers, cookies=cookies1)


        docsin = BeautifulSoup(paginasinistro.text, "html.parser")

        emailsegurado = docsin.find('input', {'name':'m_email'})['value']

        modelocarro = docsin.find('select', {'name':'m_tipo_modelo'}).string

        placasegurado = docsin.find('input', {'name': 'm_placa_ter'})['value']
        
        corretorteste = docsin.find('td', {'title': 'Clique para consultar'})['onclick']
        
        codempresa = re.findall("(?<=m_cod-empresa=).+?(?=&)", corretorteste)
        
        codsucursal = re.findall("(?<=m_cod-sucursal=).+?(?=&)", corretorteste)
        
        codcarteira = re.findall("(?<=m_cod-carteira=).+?(?=&)", corretorteste)
        
        codtipodocum = re.findall("(?<=m_tip-docum=).+?(?=&)", corretorteste)
        
        codseqdocum = re.findall("(?<=m_seq-docum=).+?(?=&)", corretorteste)
        
        codnumitem = re.findall("(?<=m_num-item=).+?(?=&)", corretorteste)
        
        codtipcorretor = re.findall("(?<=m_t_corretor=).+?(?=&)", corretorteste)
        
        tipoendosso = re.findall("(?<=m_seq-endosso=).+?(?=&)", corretorteste)
        
        codcorretor = re.findall("(?<=m_cod_corretor=).+?(?=&)", corretorteste)
        
        codopera = re.findall("(?<=m_operacao=).+?(?=&)", corretorteste)
        
        #token = re.findall("(?<=tokenSec=).+?(?=&)", corretorteste)
        
        #codhdi = re.findall("(?<=hid_chamou=).+?(?=&)", corretorteste)

        
        
        
        
        paramsapolice = { 
            "m_cod-empresa": codempresa[0],
            "m_cod-sucursal": codsucursal[0],
            "m_cod-carteira": codcarteira[0],
            "m_tip-docum": codtipodocum[0],
            "m_seq-docum": codseqdocum[0],
            "m_seq-endosso": tipoendosso[0],
            "m_num-item": codnumitem[0],
            "m_t_corretor": codtipcorretor[0],
            "m_cod_corretor": codcorretor[0], 
            "m_operacao": "consulta",
            "hid_chamou": "debt",
            "tokenSec": token,
        }  
        
        #ctrl d seleciona uma parte e replica nas outras
        
        paginaapolice = requests.post(urlapolice, data=paramsapolice, headers=Headers, cookies=cookies1)
        
        docapo = BeautifulSoup(paginaapolice.text, "html.parser")
         
        links = docapo.find_all('a', href=True)
        
        testeLink = links[0]['href']
        
        cidadeapolice = docapo.find('td', {'align': 'center'}).contents[2].strip() #indicador que escolhe a segunda opção da linha que pedi requerimento
        

        
        
        #docsin.find('input', {'name': 'm_placa_ter'})['value']
        
        
        print(cidadeapolice)
        

        
        
        
        #salva as informações na planilha              
        path = r'C:\Users\Jaime Foggiato\Desktop\ProjetosDEBT-main\planilhaEmails.xlsx'
        
        
        wb = load_workbook(path)

        ws = wb['Plan1']

        sheet = ws
        
            
        varid = sheet.cell(row=(linha+2), column=3).value

 
        ws.cell(row=(linha+2), column=1, value=1)
        ws.cell(row=(linha+2), column=4, value=cidadeapolice)
                
         

        wb.save(path)
        

    else:
        print('Já efetuado')

    linha = linha + 1
        


time.sleep(3)