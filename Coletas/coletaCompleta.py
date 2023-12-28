from cgitb import text
from email import header
from gettext import gettext
from tracemalloc import stop
from turtle import goto
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup as BS
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import csv
import time
import pandas
from openpyxl import load_workbook
from openpyxl import workbook
import requests
from bs4 import BeautifulSoup

print('Abrindo o Chrome...')
#driver = Chrome(executable_path='C:\WebDriver\bin\chromedriver.exe')
#driver = webdriver.Chrome(ChromeDriverManager().install())

driver = webdriver.Chrome()

#driver.get("https://www.hdi.com.br/hdiprestador/") talvez seja o erro


print('Acessando o Site da HDI...')
driver.get("https://www.hdi.com.br/hdiprestador/")
wait = WebDriverWait(driver,20,poll_frequency=1)

print('usuário...')
driver.find_element(By.ID, 'm_prestserv').send_keys('debt')
print('senha...')
#driver.find_element_by_id('doc').click()
driver.find_element(By.ID,'m_senha').send_keys('Agosto456*')
print('prestador...')
driver.find_element(By.XPATH, '//*[@id="m_prest_oficina"]/option[3]').click()
print('entrando')
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="login_prestador"]/div[2]/button').click()
time.sleep(1)
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="bt_pesquisar"]')))
print('loguei!')
time.sleep(1)

# como pegar as informaçõe da sessão do selenium e transformar em request
cookies1 = {}

selenium_cookies = driver.get_cookies()

for cookie in selenium_cookies:
    cookies1[cookie['name']] = cookie['value']

urlprincipal = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/dsp_cad_processo_tp2_scp.htm'

urlsinistro = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/mrsicc03x.htm'

Headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:45.0) Gecko/20100101 Firefox/45.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Charset": "ISO-8859-1,utf-8;q=0.7,*;q=0.3",
            "Accept-Encoding": "none",
            "Accept-Language": "en-US,en;q=0.8",
            "Ajax-Response": "true",
            "Connection": "keep-alive",
}


planilha = pandas.read_excel(r'C:\Users\Escritorio\Desktop\Python DEBT\infotesteHDI.xlsx', engine='openpyxl')

linha = 0


while linha != len(planilha.index):   #está correto?

    if (planilha['AUX'][linha]) == 0: #duvida nessa parte da condição

        pasta = int(planilha['PASTA'][linha])

        
        paramsprincipal = {
            "sh": "clljlkkdaidllcBd",
            "us": "debt",
            "senha": "",
            "m_us": "debt",
            "m_municipio": "",
            "m_cod_processo": pasta,
            "m_recid": "",
            "dir": "pesquisar",
            "m_chamado_ressarc_judi": "", 
            "m_chamado_ressarc_amig": "",
            "enc_us": "ZGVidA==",
        
        }

        paginaprincipal = requests.post(urlprincipal, data=paramsprincipal, headers=Headers, cookies=cookies1)

        doc = BeautifulSoup(paginaprincipal.text, "html.parser")

        valor = 'R$ ' + doc.find(id="m_vlr_base_ress")['value']

        nsinistro = doc.find(id='m_num_sinistro')['value']

        nsinistro = nsinistro[len(nsinistro)-15:]

        datasinistro = doc.find(id='m_dat_sinistro')['value']

        placadoreu = doc.find(id='m_placa_reu')['value']
            
        apolice = doc.find(id='m_num_apolice')['value']

        try:
            cpfdoreu = doc.find(id='cpf-reu')['value']
        except:
            cpfdoreu = 'CPF Não Informado'

        nomedoreu = doc.find(id='nom_reu')['value']

        congenere = doc.find(id='m_nom_congenere')['value']
            
        analista = doc. find(id='m_nom_coordenador')['value']
            
        carteira = doc.find(id='m_nom_objeto')['value']



        paramssinistro = {
            "isRevamp": "",
            "m_cor_padrao": "",
            "m_img_padrao": "",
            "m_tit_padrao": "",
            "sel_issinimport": nsinistro,
            "m_login": "debt",
            "ag_int": "0",
            "m_sinuss": nsinistro,
            "seg_cons": "" ,
            "hid_chamou": "ADVG",
            "hid_pendencia": "",
            "m_seq_envio": "",
            "cad_avaria_prv": "",
        
        }


        paginasinistro = requests.post(urlsinistro, data=paramssinistro, headers=Headers, cookies=cookies1)


        docsin = BeautifulSoup(paginasinistro.text, "html.parser")

        segurado = docsin.find('input', {'name':'m_nome_ter'})['value']

        avisosinistro = docsin.find('textarea', {'name':'m_novo'}).string
            
        veiculosegurado = docsin.find('select', {'name':'m_tipo_modelo'}).string
                
        acidenterua = docsin.find('input', {'name':'m_endereco_sin'})['value']
            
        acidentenumero = docsin.find('input', {'name':'m_numero_sin'})['value']
            
        acidentebairro = docsin.find('input', {'name':'m_bairro_sin'})['value']
                    
        acidentecidade = docsin.find('select', {'name':'m_cidade_sin'}).text
        
        acidentecidade = acidentecidade.strip()
            
        localdoacidente = acidenterua + ', ' + acidentenumero + ', ' + acidentebairro + ', ' + str(acidentecidade)
            
        try:
            tipododano = docsin.find('select', {'name':'m_compl1'}).string
        except:
            tipododano = 'Não informado'

            
        tipodoacidente = docsin.find('select', {'name':'m_causa'}).text

        emailsegurado = docsin.find('input', {'name': 'm_email'})['value']
        
            
        urlcausador = f'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/dsp_consulta_causador.htm?m_sinuss={nsinistro}'
        
        paramcausador = {
        "m_sinuss": nsinistro
    }
        
        paginacausador = requests.get(urlcausador, data=paramcausador, headers=Headers, cookies=cookies1)
        doccausador = BeautifulSoup(paginacausador.text, 'html.parser')
        
        try:
            causadorNome = doccausador.find('input', {'name':'m_nome_causador'})['value']
        except:
            causadorNome = 'Nao informado'
            
        try:
            causadorCPF = doccausador.find('input', {'name':'m_num_documento'})['value']
        except:
            causadorCPF = 'Nao informado'
            
        try:
            causadorplaca = doccausador.find('input', {'name':'m_placa_causador'})['value']
        except:
            causadorplaca = 'Nao informado'
            
        try:
            causadorTelefone = doccausador.find('input', {'name':'m_tel_causador'})['value']
        except:
            causadorTelefone = 'Nao informado'
            
        try:
            causadorEndereco = doccausador.find('input', {'name':'m_end_causador'})['value']
        except:
            causadorEndereco = 'Nao informado'
            
        try:
            causadorN =	doccausador.find('input', {'name':'num_causador'})['value']
        except:
            causadorN = '' 
            
        try:
            causadorComp = doccausador.find('input', {'name':'m_cpl_end_causador'})['value']
        except:
            causadorComp = ''
            
        try:
            causadorBairro = doccausador.find('input', {'name':'m_bairro_causador'})['value']
        except:
            causadorBairro = ''
            
        try:
            causadorCidade = doccausador.find('input', {'name':'m_cidade_causador'})['value']
        except:
            causadorCidade = ''
            
        try:
            causadorUF = doccausador.find('input', {'name':'m_uf_causador'})['value']
        except:
            causadorUF = ''
            
        try:
            causadorcep = doccausador.find('input', {'name':'m_cep_causador'})['value']
        except:
            causadorcep = ''
            
        try:
            
            causadorcongenere =	doccausador.find('input', {'name':'m_nom_congenere'})['value']
        except:
            causadorcongenere = 'Nao informado'
            
        try:
            causadorapolice = doccausador.find('input', {'name':'m_num_apolice'})['value']
        except:
            causadorapolice = 'Nao informado'
            
        try:
            causadordatainiciodevigencia = doccausador.find('input', {'name':'m_dat_inivig'})['value']
        except:
            causadordatainiciodevigencia = 'Nao informado'
            
        try:
            causadordatafinaldevigencia = doccausador.find('input', {'name':'m_dat_fimvig'})['value']
        except:
            causadordatafinaldevigencia = 'Nao informado'
            
            
        causadorenderecocompleto = causadorEndereco + ', ' +  causadorN + ', ' +  causadorComp + ', ' +  causadorBairro + ', ' +  causadorBairro + ', ' +  causadorCidade + ', ' +  causadorUF + ', ' +  causadorcep
        
            
        #salva as informações na planilha              
        path = r'C:\Users\Escritorio\Desktop\Python DEBT\infotesteHDI.xlsx' 
        
        
        wb = load_workbook(path)

        ws = wb['Plan1']

        
        ws.cell(row=(linha+2), column=1, value=1)
        ws.cell(row=(linha+2), column=3, value=valor)
        ws.cell(row=(linha+2), column=4, value=nsinistro)
        ws.cell(row=(linha+2), column=5, value=datasinistro)
        ws.cell(row=(linha+2), column=6, value=placadoreu)
        ws.cell(row=(linha+2), column=7, value=apolice)
        ws.cell(row=(linha+2), column=8, value=nomedoreu)
        ws.cell(row=(linha+2), column=9, value=cpfdoreu)
        ws.cell(row=(linha+2), column=10, value=congenere)
        ws.cell(row=(linha+2), column=11, value=segurado)
        ws.cell(row=(linha+2), column=12, value=avisosinistro)
        ws.cell(row=(linha+2), column=13, value=veiculosegurado) 
        ws.cell(row=(linha+2), column=14, value=localdoacidente) 
        ws.cell(row=(linha+2), column=15, value=tipododano) 
        ws.cell(row=(linha+2), column=16, value=tipodoacidente) 
        ws.cell(row=(linha+2), column=17, value=carteira) 
        ws.cell(row=(linha+2), column=18, value=analista)
        ws.cell(row=(linha+2), column=19, value=emailsegurado)
        ws.cell(row=(linha+2), column=20, value=str(causadorNome))
        ws.cell(row=(linha+2), column=21, value=str(causadorCPF))
        ws.cell(row=(linha+2), column=22, value=str(causadorplaca))
        ws.cell(row=(linha+2), column=23, value=str(causadorTelefone))
        ws.cell(row=(linha+2), column=24, value=str(causadorenderecocompleto))
        ws.cell(row=(linha+2), column=25, value=str(causadorcongenere))
        ws.cell(row=(linha+2), column=26, value=str(causadorapolice))
        ws.cell(row=(linha+2), column=27, value=str(causadordatainiciodevigencia))
        ws.cell(row=(linha+2), column=28, value=str(causadordatafinaldevigencia))
        #ws.cell(row=(linha+2), column=29, value=emailsegurado)
                
        

        

        wb.save(path)




        print(valor)
        print(nsinistro)
        print(datasinistro)
        print(placadoreu)
        print(apolice)
        print(cpfdoreu)
        print(nomedoreu)
        print(congenere)
        print(segurado)
        print(avisosinistro)
        print('')
        print(linha)


    else:
        print('Já efetuado')

    linha = linha + 1
        

    time.sleep(3)