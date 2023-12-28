from cgitb import text
from email import header
from gettext import gettext
from tracemalloc import stop
from turtle import goto
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup as BS
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import csv
import time
import pandas
from openpyxl import load_workbook
from openpyxl import workbook
import requests
from bs4 import BeautifulSoup
import math

print('Abrindo o Chrome...')
#driver = Chrome(executable_path='C:\WebDriver\bin\chromedriver.exe')
#driver = webdriver.Chrome(ChromeDriverManager().install())

driver = webdriver.Chrome()

#driver.get("https://www.hdi.com.br/hdiprestador/") talvez seja o erro


print('Acessando o Site da HDI...')
driver.get("https://www.hdi.com.br/hdiprestador/")
wait = WebDriverWait(driver,20,poll_frequency=1)

print('usuário...')
driver.find_element(By.ID, 'm_prestserv').send_keys('debt')
print('senha...')
#driver.find_element_by_id('doc').click()
driver.find_element(By.ID,'m_senha').send_keys('Deze963@')
print('prestador...')
driver.find_element(By.XPATH, '//*[@id="m_prest_oficina"]/option[3]').click()
print('entrando')
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="login_prestador"]/div[2]/button').click()
time.sleep(1)
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="bt_pesquisar"]')))
print('loguei!')
time.sleep(1)

# como pegar as informaçõe da sessão do selenium e transformar em request
cookies1 = {}

selenium_cookies = driver.get_cookies()

for cookie in selenium_cookies:
    cookies1[cookie['name']] = cookie['value']

urlprincipal = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/dsp_cad_processo_tp2_scp.htm'

urlsinistro = 'https://www.hdi.com.br/scripts/cgiip.exe/WService=wsbroker2/mrsicc03x.htm'

Headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:45.0) Gecko/20100101 Firefox/45.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Charset": "ISO-8859-1,utf-8;q=0.7,*;q=0.3",
            "Accept-Encoding": "none",
            "Accept-Language": "en-US,en;q=0.8",
            "Ajax-Response": "true",
            "Connection": "keep-alive",
}



planilha = pandas.read_excel(r'C:\Users\Jaime Foggiato\Desktop\Projetos Debt\Planilhas Segurado.xlsx', engine='openpyxl') #não esquecer de mudar o diretório, toda vez que for rodar em um pc novo.

linha = 0


while linha != len(planilha.index):  

    if (planilha['AUX'][linha]) == 0: 

        pasta = int(planilha['PASTA'][linha])

        
        paramsprincipal = {
            "sh": "clljlkkdaidllcBd",
            "us": "debt",
            "senha": "",
            "m_us": "debt",
            "m_municipio": "",
            "m_cod_processo": pasta,
            "m_recid": "",
            "dir": "pesquisar",
            "m_chamado_ressarc_judi": "", 
            "m_chamado_ressarc_amig": "",
            "enc_us": "ZGVidA==",
        
        }

        paginaprincipal = requests.post(urlprincipal, data=paramsprincipal, headers=Headers, cookies=cookies1)

        doc = BeautifulSoup(paginaprincipal.text, "html.parser")

        nsinistro_element = doc.find(id='m_num_sinistro')
        
        if nsinistro_element:
            nsinistro = nsinistro_element.get('value', 'N/A')
        else:
            nsinistro = 'Elemento não encontrado'
            linha += 1
            continue

        nsinistro = nsinistro[len(nsinistro)-15:]

        datasinistro = doc.find(id='m_dat_sinistro')['value']



        paramssinistro = {
            "isRevamp": "",
            "m_cor_padrao": "",
            "m_img_padrao": "",
            "m_tit_padrao": "",
            "sel_issinimport": nsinistro,
            "m_login": "debt",
            "ag_int": "0",
            "m_sinuss": nsinistro,
            "seg_cons": "" ,
            "hid_chamou": "ADVG",
            "hid_pendencia": "",
            "m_seq_envio": "",
            "cad_avaria_prv": "",
        
        }


        paginasinistro = requests.post(urlsinistro, data=paramssinistro, headers=Headers, cookies=cookies1)


        docsin = BeautifulSoup(paginasinistro.text, "html.parser")

        nomesegurado = docsin.find('input', {'name': 'm_nome_ter'})['value']
        
        telefonesegurado = docsin.find_all('input', {'name': 'm_fone_ter'})
        
        if telefonesegurado:
            telefonesegurado = [telefone['value'] for telefone in telefonesegurado]
        else:
            telefonesegurado = ['Elemento não encontrado']

        celularsegurado = telefonesegurado[0]
        if len(telefonesegurado) > 1:
            celularsegurado = telefonesegurado[1]
            
        

        

        
            
        #salva as informações na planilha              

        path = r'C:\Users\Jaime Foggiato\Desktop\Projetos Debt\Planilhas Segurado.xlsx'

        
        
        wb = load_workbook(path)

        ws = wb['Plan1']

        sheet = ws
        
            
        varid = sheet.cell(row=(linha+2), column=3).value


        
        ws.cell(row=(linha+2), column=1, value=1)
        ws.cell(row=(linha+2), column=4, value=nomesegurado)
        ws.cell(row=(linha+2), column=5, value=nsinistro)
        ws.cell(row=(linha+2), column=6, value=', '.join(telefonesegurado))
        ws.cell(row=(linha+2), column=7, value=celularsegurado)
        
        if telefonesegurado:
            ws.cell(row=(linha+2), column=6, value=', '.join(telefonesegurado))
        else:
            ws.cell(row=(linha+2), column=6, value='Elemento não encontrado')
            ws.cell(row=(linha+2), column=7, value=celularsegurado)
            
        ws.cell(row=(linha+2), column=8, value=nsinistro)
        
              
         

        wb.save(path)



        print(nsinistro)
        print(datasinistro)
        print('')
        print(linha)
        print(telefonesegurado)
        print(celularsegurado)


    else:
        print('Já efetuado')

    linha = linha + 1
        

    time.sleep(3)

# Read the Excel file into a DataFrame
df = pandas.read_excel(r'C:\Users\Jaime Foggiato\Desktop\Projetos Debt\Planilhas Segurado.xlsx')
